"""Repository for all Reports & Exports queries."""

from __future__ import annotations

import csv
import io
import math
import uuid
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

from sqlalchemy import func, or_, select, text
from sqlalchemy.orm import Session, aliased

from app.models.customer import Customer, CustomerStatus, CustomerType
from app.models.invoice import Invoice, InvoiceStatus
from app.models.payment import Payment, PaymentMethod
from app.models.subscription import Subscription, SubscriptionStatus


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _apply_invoice_revenue_filters(stmt, *, date_from=None, date_to=None, plan=None, customer_type=None):
    stmt = stmt.where(Invoice.status.not_in([InvoiceStatus.DRAFT, InvoiceStatus.CANCELLED]))
    if date_from:
        stmt = stmt.where(Invoice.invoice_date >= date_from)
    if date_to:
        stmt = stmt.where(Invoice.invoice_date <= date_to)
    if plan:
        stmt = stmt.where(Invoice.plan_name_snapshot.ilike(f"%{plan}%"))
    return stmt


# ---------------------------------------------------------------------------
# Main Repository
# ---------------------------------------------------------------------------

class ReportsRepository:
    def __init__(self, db: Session) -> None:
        self.db = db

    # ── Customer Report ────────────────────────────────────────────────────

    def customers_report(
        self,
        *,
        page: int = 1,
        page_size: int = 25,
        search: str = "",
        sort_by: str = "created_at",
        sort_order: str = "desc",
        status_filter: str | None = None,
        customer_type_filter: str | None = None,
        city_filter: str | None = None,
        reference_source_filter: str | None = None,
        sales_person_filter: str | None = None,
        date_from: date | None = None,
        date_to: date | None = None,
        for_export: bool = False,
    ) -> tuple[list[dict], int, dict]:
        # Correlated subquery: active subscription count
        active_sub_sq = (
            select(func.count(Subscription.id))
            .where(Subscription.customer_id == Customer.id)
            .where(Subscription.status == SubscriptionStatus.ACTIVE)
            .where(Subscription.deleted_at.is_(None))
            .correlate(Customer)
            .scalar_subquery()
        )

        # Correlated subquery: outstanding via subscription path
        outstanding_sub_sq = (
            select(func.coalesce(func.sum(Invoice.balance_amount), 0))
            .join(Subscription, Invoice.subscription_id == Subscription.id)
            .where(Subscription.customer_id == Customer.id)
            .where(Subscription.deleted_at.is_(None))
            .where(Invoice.deleted_at.is_(None))
            .where(Invoice.balance_amount > 0)
            .correlate(Customer)
            .scalar_subquery()
        )

        # Correlated subquery: outstanding via direct customer_id (CONSOLIDATED invoices)
        outstanding_direct_sq = (
            select(func.coalesce(func.sum(Invoice.balance_amount), 0))
            .where(Invoice.customer_id == Customer.id)
            .where(Invoice.deleted_at.is_(None))
            .where(Invoice.balance_amount > 0)
            .correlate(Customer)
            .scalar_subquery()
        )

        stmt = (
            select(
                Customer.id,
                Customer.customer_code,
                Customer.full_name,
                Customer.customer_type,
                Customer.city,
                Customer.mobile_number,
                Customer.status,
                Customer.reference_source,
                Customer.sales_person,
                Customer.created_at,
                active_sub_sq.label("active_subscription_count"),
                (outstanding_sub_sq + outstanding_direct_sq).label("outstanding_amount"),
            )
            .where(Customer.deleted_at.is_(None))
        )

        if search:
            term = f"%{search}%"
            stmt = stmt.where(
                or_(
                    Customer.customer_code.ilike(term),
                    Customer.full_name.ilike(term),
                    Customer.mobile_number.ilike(term),
                    Customer.email.ilike(term),
                )
            )
        if status_filter:
            try:
                stmt = stmt.where(Customer.status == CustomerStatus(status_filter))
            except ValueError:
                pass
        if customer_type_filter:
            try:
                stmt = stmt.where(Customer.customer_type == CustomerType(customer_type_filter))
            except ValueError:
                pass
        if city_filter:
            stmt = stmt.where(Customer.city.ilike(f"%{city_filter}%"))
        if reference_source_filter:
            stmt = stmt.where(Customer.reference_source.ilike(f"%{reference_source_filter}%"))
        if sales_person_filter:
            stmt = stmt.where(Customer.sales_person.ilike(f"%{sales_person_filter}%"))
        if date_from:
            stmt = stmt.where(Customer.created_at >= date_from)
        if date_to:
            stmt = stmt.where(Customer.created_at <= date_to)

        _sort_map = {
            "customer_code": Customer.customer_code,
            "full_name": Customer.full_name,
            "status": Customer.status,
            "created_at": Customer.created_at,
            "city": Customer.city,
        }
        col = _sort_map.get(sort_by, Customer.created_at)
        stmt = stmt.order_by(col.desc() if sort_order == "desc" else col.asc())

        total: int = self.db.scalar(select(func.count()).select_from(stmt.subquery())) or 0

        if for_export:
            rows = self.db.execute(stmt).all()
        else:
            rows = self.db.execute(stmt.offset((page - 1) * page_size).limit(page_size)).all()

        items = [
            {
                "id": str(r.id),
                "customer_code": r.customer_code,
                "full_name": r.full_name,
                "customer_type": r.customer_type,
                "city": r.city or "",
                "mobile_number": r.mobile_number,
                "active_subscription_count": r.active_subscription_count or 0,
                "outstanding_amount": float(r.outstanding_amount or 0),
                "status": r.status,
            }
            for r in rows
        ]

        # Summary
        summary_rows = self.db.execute(
            select(
                func.count(Customer.id).label("total"),
                func.count(Customer.id).filter(Customer.status == CustomerStatus.ACTIVE).label("active"),
                func.count(Customer.id).filter(Customer.customer_type == CustomerType.BUSINESS).label("business"),
                func.count(Customer.id).filter(Customer.customer_type == CustomerType.INDIVIDUAL).label("individual"),
            ).where(Customer.deleted_at.is_(None))
        ).one()
        summary = {
            "total_customers": summary_rows.total,
            "active_customers": summary_rows.active,
            "business_customers": summary_rows.business,
            "individual_customers": summary_rows.individual,
        }
        return items, total, summary

    # ── Subscription Report ────────────────────────────────────────────────

    def subscriptions_report(
        self,
        *,
        page: int = 1,
        page_size: int = 25,
        search: str = "",
        sort_by: str = "expiry_date",
        sort_order: str = "asc",
        status_filter: str | None = None,
        customer_filter: str | None = None,
        plan_filter: str | None = None,
        sub_date_from: date | None = None,
        sub_date_to: date | None = None,
        expiry_date_from: date | None = None,
        expiry_date_to: date | None = None,
        quick_filter: str | None = None,
        for_export: bool = False,
    ) -> tuple[list[dict], int, dict]:
        stmt = (
            select(
                Subscription.id,
                Subscription.subscription_code,
                Customer.full_name.label("customer_name"),
                Customer.customer_code.label("customer_code"),
                Subscription.connection_name,
                Subscription.plan_name_snapshot,
                Subscription.billing_cycle_snapshot,
                Subscription.start_date,
                Subscription.renewal_date,
                Subscription.expiry_date,
                Subscription.status,
            )
            .join(Customer, Subscription.customer_id == Customer.id)
            .where(Subscription.deleted_at.is_(None))
            .where(Customer.deleted_at.is_(None))
        )

        if search:
            term = f"%{search}%"
            stmt = stmt.where(
                or_(
                    Subscription.subscription_code.ilike(term),
                    Customer.full_name.ilike(term),
                    Subscription.connection_name.ilike(term),
                )
            )
        if status_filter:
            try:
                stmt = stmt.where(Subscription.status == SubscriptionStatus(status_filter))
            except ValueError:
                pass
        if customer_filter:
            term = f"%{customer_filter}%"
            stmt = stmt.where(
                or_(Customer.full_name.ilike(term), Customer.customer_code.ilike(term))
            )
        if plan_filter:
            stmt = stmt.where(Subscription.plan_name_snapshot.ilike(f"%{plan_filter}%"))
        if sub_date_from:
            stmt = stmt.where(Subscription.start_date >= sub_date_from)
        if sub_date_to:
            stmt = stmt.where(Subscription.start_date <= sub_date_to)
        if expiry_date_from:
            stmt = stmt.where(Subscription.expiry_date >= expiry_date_from)
        if expiry_date_to:
            stmt = stmt.where(Subscription.expiry_date <= expiry_date_to)

        today = date.today()
        if quick_filter == "7d":
            stmt = stmt.where(Subscription.expiry_date >= today)
            stmt = stmt.where(Subscription.expiry_date <= today + timedelta(days=7))
            stmt = stmt.where(Subscription.status == SubscriptionStatus.ACTIVE)
        elif quick_filter == "15d":
            stmt = stmt.where(Subscription.expiry_date >= today)
            stmt = stmt.where(Subscription.expiry_date <= today + timedelta(days=15))
            stmt = stmt.where(Subscription.status == SubscriptionStatus.ACTIVE)
        elif quick_filter == "30d":
            stmt = stmt.where(Subscription.expiry_date >= today)
            stmt = stmt.where(Subscription.expiry_date <= today + timedelta(days=30))
            stmt = stmt.where(Subscription.status == SubscriptionStatus.ACTIVE)
        elif quick_filter == "expired":
            stmt = stmt.where(Subscription.status == SubscriptionStatus.EXPIRED)

        _sort_map = {
            "subscription_code": Subscription.subscription_code,
            "customer_name": Customer.full_name,
            "expiry_date": Subscription.expiry_date,
            "start_date": Subscription.start_date,
            "renewal_date": Subscription.renewal_date,
            "status": Subscription.status,
        }
        col = _sort_map.get(sort_by, Subscription.expiry_date)
        stmt = stmt.order_by(col.desc() if sort_order == "desc" else col.asc())

        total: int = self.db.scalar(select(func.count()).select_from(stmt.subquery())) or 0

        if for_export:
            rows = self.db.execute(stmt).all()
        else:
            rows = self.db.execute(stmt.offset((page - 1) * page_size).limit(page_size)).all()

        items = [
            {
                "id": str(r.id),
                "subscription_code": r.subscription_code,
                "customer_name": r.customer_name,
                "customer_code": r.customer_code,
                "connection_name": r.connection_name or "",
                "plan_name": r.plan_name_snapshot,
                "billing_cycle": r.billing_cycle_snapshot,
                "start_date": r.start_date.isoformat(),
                "renewal_date": r.renewal_date.isoformat(),
                "expiry_date": r.expiry_date.isoformat(),
                "status": r.status,
            }
            for r in rows
        ]

        summary_rows = self.db.execute(
            select(
                func.count(Subscription.id).label("total"),
                func.count(Subscription.id).filter(Subscription.status == SubscriptionStatus.ACTIVE).label("active"),
                func.count(Subscription.id).filter(
                    Subscription.status == SubscriptionStatus.ACTIVE,
                    Subscription.expiry_date >= today,
                    Subscription.expiry_date <= today + timedelta(days=30),
                ).label("expiring_soon"),
                func.count(Subscription.id).filter(Subscription.status == SubscriptionStatus.EXPIRED).label("expired"),
            ).where(Subscription.deleted_at.is_(None))
        ).one()
        summary = {
            "total_subscriptions": summary_rows.total,
            "active_subscriptions": summary_rows.active,
            "expiring_soon": summary_rows.expiring_soon,
            "expired": summary_rows.expired,
        }
        return items, total, summary

    # ── Invoice Report ─────────────────────────────────────────────────────

    def invoices_report(
        self,
        *,
        page: int = 1,
        page_size: int = 25,
        search: str = "",
        sort_by: str = "created_at",
        sort_order: str = "desc",
        status_filter: str | None = None,
        customer_filter: str | None = None,
        plan_filter: str | None = None,
        invoice_date_from: date | None = None,
        invoice_date_to: date | None = None,
        due_date_from: date | None = None,
        due_date_to: date | None = None,
        quick_filter: str | None = None,
        for_export: bool = False,
    ) -> tuple[list[dict], int, dict]:
        stmt = (
            select(
                Invoice.id,
                Invoice.invoice_number,
                Invoice.customer_name_snapshot,
                Invoice.connection_name_snapshot,
                Invoice.plan_name_snapshot,
                Invoice.invoice_date,
                Invoice.due_date,
                Invoice.total_amount,
                Invoice.paid_amount,
                Invoice.balance_amount,
                Invoice.status,
            )
            .where(Invoice.deleted_at.is_(None))
        )

        if search:
            term = f"%{search}%"
            stmt = stmt.where(
                or_(
                    Invoice.invoice_number.ilike(term),
                    Invoice.customer_name_snapshot.ilike(term),
                    Invoice.customer_code_snapshot.ilike(term),
                    Invoice.connection_name_snapshot.ilike(term),
                )
            )
        if status_filter:
            try:
                stmt = stmt.where(Invoice.status == InvoiceStatus(status_filter))
            except ValueError:
                pass
        if customer_filter:
            term = f"%{customer_filter}%"
            stmt = stmt.where(
                or_(
                    Invoice.customer_name_snapshot.ilike(term),
                    Invoice.customer_code_snapshot.ilike(term),
                )
            )
        if plan_filter:
            stmt = stmt.where(Invoice.plan_name_snapshot.ilike(f"%{plan_filter}%"))
        if invoice_date_from:
            stmt = stmt.where(Invoice.invoice_date >= invoice_date_from)
        if invoice_date_to:
            stmt = stmt.where(Invoice.invoice_date <= invoice_date_to)
        if due_date_from:
            stmt = stmt.where(Invoice.due_date >= due_date_from)
        if due_date_to:
            stmt = stmt.where(Invoice.due_date <= due_date_to)

        today = date.today()
        if quick_filter == "due_today":
            stmt = stmt.where(Invoice.due_date == today)
            stmt = stmt.where(Invoice.balance_amount > 0)
        elif quick_filter == "due_7d":
            stmt = stmt.where(Invoice.due_date >= today)
            stmt = stmt.where(Invoice.due_date <= today + timedelta(days=7))
            stmt = stmt.where(Invoice.balance_amount > 0)
        elif quick_filter == "due_15d":
            stmt = stmt.where(Invoice.due_date >= today)
            stmt = stmt.where(Invoice.due_date <= today + timedelta(days=15))
            stmt = stmt.where(Invoice.balance_amount > 0)
        elif quick_filter == "due_30d":
            stmt = stmt.where(Invoice.due_date >= today)
            stmt = stmt.where(Invoice.due_date <= today + timedelta(days=30))
            stmt = stmt.where(Invoice.balance_amount > 0)
        elif quick_filter == "overdue":
            stmt = stmt.where(Invoice.due_date < today)
            stmt = stmt.where(Invoice.balance_amount > 0)

        _sort_map = {
            "invoice_date": Invoice.invoice_date,
            "due_date": Invoice.due_date,
            "total_amount": Invoice.total_amount,
            "balance_amount": Invoice.balance_amount,
            "status": Invoice.status,
            "created_at": Invoice.created_at,
            "invoice_number": Invoice.invoice_number,
        }
        col = _sort_map.get(sort_by, Invoice.created_at)
        stmt = stmt.order_by(col.desc() if sort_order == "desc" else col.asc())

        total: int = self.db.scalar(select(func.count()).select_from(stmt.subquery())) or 0

        if for_export:
            rows = self.db.execute(stmt).all()
        else:
            rows = self.db.execute(stmt.offset((page - 1) * page_size).limit(page_size)).all()

        items = [
            {
                "id": str(r.id),
                "invoice_number": r.invoice_number,
                "customer_name": r.customer_name_snapshot,
                "connection_name": r.connection_name_snapshot or "",
                "plan_name": r.plan_name_snapshot or "",
                "invoice_date": r.invoice_date.isoformat() if r.invoice_date else "",
                "due_date": r.due_date.isoformat() if r.due_date else "",
                "total_amount": float(r.total_amount or 0),
                "paid_amount": float(r.paid_amount or 0),
                "balance_amount": float(r.balance_amount or 0),
                "status": r.status,
            }
            for r in rows
        ]

        summary_rows = self.db.execute(
            select(
                func.count(Invoice.id).label("total"),
                func.coalesce(func.sum(Invoice.total_amount), 0).label("total_invoiced"),
                func.coalesce(func.sum(Invoice.paid_amount), 0).label("total_collected"),
                func.coalesce(func.sum(Invoice.balance_amount), 0).label("total_outstanding"),
            ).where(Invoice.deleted_at.is_(None))
        ).one()
        summary = {
            "total_invoices": summary_rows.total,
            "total_invoiced_amount": float(summary_rows.total_invoiced),
            "total_collected_amount": float(summary_rows.total_collected),
            "total_outstanding_amount": float(summary_rows.total_outstanding),
        }
        return items, total, summary

    # ── Payment Report ─────────────────────────────────────────────────────

    def payments_report(
        self,
        *,
        page: int = 1,
        page_size: int = 25,
        search: str = "",
        sort_by: str = "payment_date",
        sort_order: str = "desc",
        customer_filter: str | None = None,
        payment_method_filter: str | None = None,
        date_from: date | None = None,
        date_to: date | None = None,
        for_export: bool = False,
    ) -> tuple[list[dict], int, dict]:
        stmt = (
            select(
                Payment.id,
                Payment.payment_number,
                Invoice.invoice_number,
                Invoice.customer_name_snapshot,
                Payment.amount,
                Payment.payment_method,
                Payment.payment_date,
                Payment.transaction_reference,
            )
            .join(Invoice, Payment.invoice_id == Invoice.id)
            .where(Payment.deleted_at.is_(None))
        )

        if search:
            term = f"%{search}%"
            stmt = stmt.where(
                or_(
                    Payment.payment_number.ilike(term),
                    Invoice.invoice_number.ilike(term),
                    Payment.transaction_reference.ilike(term),
                )
            )
        if customer_filter:
            term = f"%{customer_filter}%"
            stmt = stmt.where(Invoice.customer_name_snapshot.ilike(term))
        if payment_method_filter:
            try:
                stmt = stmt.where(Payment.payment_method == PaymentMethod(payment_method_filter))
            except ValueError:
                pass
        if date_from:
            stmt = stmt.where(Payment.payment_date >= date_from)
        if date_to:
            stmt = stmt.where(Payment.payment_date <= date_to)

        _sort_map = {
            "payment_date": Payment.payment_date,
            "amount": Payment.amount,
            "payment_number": Payment.payment_number,
        }
        col = _sort_map.get(sort_by, Payment.payment_date)
        stmt = stmt.order_by(col.desc() if sort_order == "desc" else col.asc())

        total: int = self.db.scalar(select(func.count()).select_from(stmt.subquery())) or 0

        if for_export:
            rows = self.db.execute(stmt).all()
        else:
            rows = self.db.execute(stmt.offset((page - 1) * page_size).limit(page_size)).all()

        items = [
            {
                "id": str(r.id),
                "payment_number": r.payment_number,
                "invoice_number": r.invoice_number,
                "customer_name": r.customer_name_snapshot,
                "amount": float(r.amount),
                "payment_method": r.payment_method,
                "payment_date": r.payment_date.isoformat(),
                "transaction_reference": r.transaction_reference or "",
            }
            for r in rows
        ]

        summary_rows = self.db.execute(
            select(
                func.count(Payment.id).label("total"),
                func.coalesce(func.sum(Payment.amount), 0).label("total_amount"),
            ).where(Payment.deleted_at.is_(None))
        ).one()
        summary = {
            "total_payments": summary_rows.total,
            "total_collection_amount": float(summary_rows.total_amount),
        }
        return items, total, summary

    # ── Revenue Report ─────────────────────────────────────────────────────

    def revenue_report(
        self,
        *,
        date_from: date | None = None,
        date_to: date | None = None,
        plan_filter: str | None = None,
        customer_type_filter: str | None = None,
        city_filter: str | None = None,
    ) -> dict:
        base_where = [
            Invoice.deleted_at.is_(None),
            Invoice.status.not_in([InvoiceStatus.DRAFT, InvoiceStatus.CANCELLED]),
        ]
        if date_from:
            base_where.append(Invoice.invoice_date >= date_from)
        if date_to:
            base_where.append(Invoice.invoice_date <= date_to)
        if plan_filter:
            base_where.append(Invoice.plan_name_snapshot.ilike(f"%{plan_filter}%"))

        # Revenue by month
        by_month_rows = self.db.execute(
            select(
                func.date_trunc("month", Invoice.invoice_date).label("month"),
                func.sum(Invoice.total_amount).label("revenue"),
            )
            .where(*base_where)
            .group_by(func.date_trunc("month", Invoice.invoice_date))
            .order_by(func.date_trunc("month", Invoice.invoice_date))
            .limit(24)
        ).all()

        revenue_by_month = [
            {
                "month": r.month.strftime("%Y-%m"),
                "label": r.month.strftime("%b %Y"),
                "revenue": float(r.revenue or 0),
            }
            for r in by_month_rows
        ]

        # Revenue by plan
        by_plan_rows = self.db.execute(
            select(
                Invoice.plan_name_snapshot.label("plan_name"),
                func.sum(Invoice.total_amount).label("revenue"),
            )
            .where(*base_where)
            .group_by(Invoice.plan_name_snapshot)
            .order_by(func.sum(Invoice.total_amount).desc())
            .limit(10)
        ).all()

        revenue_by_plan = [
            {"plan_name": r.plan_name, "revenue": float(r.revenue or 0)}
            for r in by_plan_rows
        ]

        # Revenue by customer (top 10)
        by_customer_rows = self.db.execute(
            select(
                Invoice.customer_name_snapshot.label("customer_name"),
                func.sum(Invoice.total_amount).label("revenue"),
            )
            .where(*base_where)
            .group_by(Invoice.customer_name_snapshot)
            .order_by(func.sum(Invoice.total_amount).desc())
            .limit(10)
        ).all()

        revenue_by_customer = [
            {"customer_name": r.customer_name, "revenue": float(r.revenue or 0)}
            for r in by_customer_rows
        ]

        # Revenue by city — join through subscription→customer for SINGLE, direct for CONSOLIDATED
        c1 = aliased(Customer)
        c2 = aliased(Customer)
        city_base = (
            select(
                func.coalesce(c1.city, c2.city, "Unknown").label("city"),
                func.sum(Invoice.total_amount).label("revenue"),
            )
            .outerjoin(Subscription, Invoice.subscription_id == Subscription.id)
            .outerjoin(c1, Subscription.customer_id == c1.id)
            .outerjoin(c2, Invoice.customer_id == c2.id)
            .where(Invoice.deleted_at.is_(None))
            .where(Invoice.status.not_in([InvoiceStatus.DRAFT, InvoiceStatus.CANCELLED]))
        )
        if date_from:
            city_base = city_base.where(Invoice.invoice_date >= date_from)
        if date_to:
            city_base = city_base.where(Invoice.invoice_date <= date_to)
        if city_filter:
            city_base = city_base.where(
                or_(c1.city.ilike(f"%{city_filter}%"), c2.city.ilike(f"%{city_filter}%"))
            )
        if customer_type_filter:
            try:
                ct = CustomerType(customer_type_filter)
                city_base = city_base.where(
                    or_(c1.customer_type == ct, c2.customer_type == ct)
                )
            except ValueError:
                pass

        by_city_rows = self.db.execute(
            city_base
            .group_by(func.coalesce(c1.city, c2.city, "Unknown"))
            .order_by(func.sum(Invoice.total_amount).desc())
            .limit(10)
        ).all()

        revenue_by_city = [
            {"city": r.city or "Unknown", "revenue": float(r.revenue or 0)}
            for r in by_city_rows
        ]

        # Summary
        summary_row = self.db.execute(
            select(
                func.coalesce(func.sum(Invoice.total_amount), 0).label("total_revenue"),
                func.count(func.distinct(Invoice.customer_name_snapshot)).label("unique_customers"),
                func.count(Invoice.id).filter(Invoice.subscription_id.is_not(None)).label("sub_invoices"),
            ).where(*base_where)
        ).one()

        total_rev = float(summary_row.total_revenue or 0)
        unique_customers = int(summary_row.unique_customers or 1)
        sub_invoices = int(summary_row.sub_invoices or 1)

        return {
            "revenue_by_month": revenue_by_month,
            "revenue_by_plan": revenue_by_plan,
            "revenue_by_customer": revenue_by_customer,
            "revenue_by_city": revenue_by_city,
            "summary": {
                "total_revenue": total_rev,
                "avg_revenue_per_customer": round(total_rev / unique_customers, 2) if unique_customers else 0,
                "avg_revenue_per_subscription": round(total_rev / sub_invoices, 2) if sub_invoices else 0,
            },
        }

    # ── Outstanding Report ─────────────────────────────────────────────────

    def outstanding_report(
        self,
        *,
        page: int = 1,
        page_size: int = 25,
        search: str = "",
        sort_by: str = "days_overdue",
        sort_order: str = "desc",
        customer_filter: str | None = None,
        city_filter: str | None = None,
        plan_filter: str | None = None,
        for_export: bool = False,
    ) -> tuple[list[dict], int, dict]:
        today = date.today()

        stmt = (
            select(
                Invoice.id,
                Invoice.invoice_number,
                Invoice.customer_name_snapshot,
                Invoice.connection_name_snapshot,
                Invoice.due_date,
                Invoice.balance_amount,
                Invoice.status,
                (func.current_date() - Invoice.due_date).label("days_overdue"),
            )
            .where(Invoice.deleted_at.is_(None))
            .where(Invoice.balance_amount > 0)
        )

        if search:
            term = f"%{search}%"
            stmt = stmt.where(
                or_(
                    Invoice.customer_name_snapshot.ilike(term),
                    Invoice.invoice_number.ilike(term),
                )
            )
        if customer_filter:
            stmt = stmt.where(Invoice.customer_name_snapshot.ilike(f"%{customer_filter}%"))
        if plan_filter:
            stmt = stmt.where(Invoice.plan_name_snapshot.ilike(f"%{plan_filter}%"))

        # City filter requires a join
        if city_filter:
            c1 = aliased(Customer)
            c2 = aliased(Customer)
            stmt = (
                stmt
                .outerjoin(Subscription, Invoice.subscription_id == Subscription.id)
                .outerjoin(c1, Subscription.customer_id == c1.id)
                .outerjoin(c2, Invoice.customer_id == c2.id)
                .where(
                    or_(
                        c1.city.ilike(f"%{city_filter}%"),
                        c2.city.ilike(f"%{city_filter}%"),
                    )
                )
            )

        # Sorting — days_overdue is computed; use due_date as proxy
        _sort_map = {
            "days_overdue": Invoice.due_date,
            "due_date": Invoice.due_date,
            "outstanding_amount": Invoice.balance_amount,
            "invoice_number": Invoice.invoice_number,
            "customer_name": Invoice.customer_name_snapshot,
        }
        col = _sort_map.get(sort_by, Invoice.due_date)
        # Flipping: days_overdue desc = due_date asc
        if sort_by == "days_overdue":
            stmt = stmt.order_by(col.asc() if sort_order == "desc" else col.desc())
        else:
            stmt = stmt.order_by(col.desc() if sort_order == "desc" else col.asc())

        total: int = self.db.scalar(select(func.count()).select_from(stmt.subquery())) or 0

        if for_export:
            rows = self.db.execute(stmt).all()
        else:
            rows = self.db.execute(stmt.offset((page - 1) * page_size).limit(page_size)).all()

        def _aging_bucket(days: int) -> str:
            if days <= 0:
                return "Current"
            elif days <= 30:
                return "0-30"
            elif days <= 60:
                return "31-60"
            elif days <= 90:
                return "61-90"
            return "90+"

        items = [
            {
                "id": str(r.id),
                "invoice_number": r.invoice_number,
                "customer_name": r.customer_name_snapshot,
                "connection_name": r.connection_name_snapshot or "",
                "due_date": r.due_date.isoformat(),
                "outstanding_amount": float(r.balance_amount or 0),
                "days_overdue": int(r.days_overdue or 0),
                "aging_bucket": _aging_bucket(int(r.days_overdue or 0)),
                "status": r.status.value if hasattr(r.status, "value") else str(r.status),
            }
            for r in rows
        ]

        # Aging bucket summary
        summary_row = self.db.execute(
            select(
                func.coalesce(func.sum(Invoice.balance_amount), 0).label("total_outstanding"),
                func.coalesce(
                    func.sum(Invoice.balance_amount).filter(
                        func.current_date() - Invoice.due_date <= 0
                    ), 0
                ).label("bucket_current"),
                func.coalesce(
                    func.sum(Invoice.balance_amount).filter(
                        (func.current_date() - Invoice.due_date).between(1, 30)
                    ), 0
                ).label("bucket_0_30"),
                func.coalesce(
                    func.sum(Invoice.balance_amount).filter(
                        (func.current_date() - Invoice.due_date).between(31, 60)
                    ), 0
                ).label("bucket_31_60"),
                func.coalesce(
                    func.sum(Invoice.balance_amount).filter(
                        (func.current_date() - Invoice.due_date).between(61, 90)
                    ), 0
                ).label("bucket_61_90"),
                func.coalesce(
                    func.sum(Invoice.balance_amount).filter(
                        func.current_date() - Invoice.due_date > 90
                    ), 0
                ).label("bucket_90_plus"),
            )
            .where(Invoice.deleted_at.is_(None))
            .where(Invoice.balance_amount > 0)
        ).one()

        summary = {
            "total_outstanding": float(summary_row.total_outstanding),
            "bucket_current": float(summary_row.bucket_current),
            "bucket_0_30": float(summary_row.bucket_0_30),
            "bucket_31_60": float(summary_row.bucket_31_60),
            "bucket_61_90": float(summary_row.bucket_61_90),
            "bucket_90_plus": float(summary_row.bucket_90_plus),
        }
        return items, total, summary

    # ── Export ─────────────────────────────────────────────────────────────

    def generate_export(
        self,
        report_type: str,
        filters: dict,
        fmt: str,
        storage_root: str,
    ) -> tuple[str, datetime]:
        """Generate CSV or XLSX export. Returns (file_path, expires_at)."""
        items, _, _ = self._get_export_data(report_type, filters)
        if not items:
            items = []

        timestamp = datetime.now(tz=timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"{report_type}_{timestamp}_{uuid.uuid4().hex[:8]}.{fmt}"
        exports_dir = Path(storage_root) / "exports"
        exports_dir.mkdir(parents=True, exist_ok=True)

        file_path = exports_dir / filename

        if fmt == "csv":
            self._write_csv(items, file_path)
        else:
            self._write_xlsx(items, file_path, report_type)

        # Cleanup files older than 24 hours
        self._cleanup_old_exports(exports_dir)

        expires_at = datetime.now(tz=timezone.utc) + timedelta(hours=24)
        return filename, expires_at

    def _get_export_data(self, report_type: str, filters: dict) -> tuple[list[dict], int, dict]:
        _parse_date = lambda v: date.fromisoformat(v) if v else None

        if report_type == "customers":
            return self.customers_report(
                for_export=True,
                search=filters.get("search", ""),
                status_filter=filters.get("status"),
                customer_type_filter=filters.get("customer_type"),
                city_filter=filters.get("city"),
                reference_source_filter=filters.get("reference_source"),
                sales_person_filter=filters.get("sales_person"),
                date_from=_parse_date(filters.get("date_from")),
                date_to=_parse_date(filters.get("date_to")),
            )
        elif report_type == "subscriptions":
            return self.subscriptions_report(
                for_export=True,
                search=filters.get("search", ""),
                status_filter=filters.get("status"),
                customer_filter=filters.get("customer"),
                plan_filter=filters.get("plan"),
                sub_date_from=_parse_date(filters.get("sub_date_from")),
                sub_date_to=_parse_date(filters.get("sub_date_to")),
                expiry_date_from=_parse_date(filters.get("expiry_date_from")),
                expiry_date_to=_parse_date(filters.get("expiry_date_to")),
                quick_filter=filters.get("quick_filter"),
            )
        elif report_type == "invoices":
            return self.invoices_report(
                for_export=True,
                search=filters.get("search", ""),
                status_filter=filters.get("status"),
                customer_filter=filters.get("customer"),
                plan_filter=filters.get("plan"),
                invoice_date_from=_parse_date(filters.get("invoice_date_from")),
                invoice_date_to=_parse_date(filters.get("invoice_date_to")),
                due_date_from=_parse_date(filters.get("due_date_from")),
                due_date_to=_parse_date(filters.get("due_date_to")),
                quick_filter=filters.get("quick_filter"),
            )
        elif report_type == "payments":
            return self.payments_report(
                for_export=True,
                search=filters.get("search", ""),
                customer_filter=filters.get("customer"),
                payment_method_filter=filters.get("payment_method"),
                date_from=_parse_date(filters.get("date_from")),
                date_to=_parse_date(filters.get("date_to")),
            )
        elif report_type == "outstanding":
            return self.outstanding_report(
                for_export=True,
                search=filters.get("search", ""),
                customer_filter=filters.get("customer"),
                city_filter=filters.get("city"),
                plan_filter=filters.get("plan"),
            )
        return [], 0, {}

    def _write_csv(self, items: list[dict], file_path: Path) -> None:
        if not items:
            file_path.write_text("No data")
            return
        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=items[0].keys())
            writer.writeheader()
            writer.writerows(items)

    def _write_xlsx(self, items: list[dict], file_path: Path, report_type: str) -> None:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise RuntimeError("openpyxl is required for XLSX export")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_type.capitalize()

        header_fill = PatternFill("solid", fgColor="1F4959")
        header_font = Font(bold=True, color="FFFFFF")

        if items:
            headers = list(items[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row_idx, item in enumerate(items, 2):
                for col_idx, key in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=item.get(key, ""))

            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col) + 4
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 40)

        wb.save(file_path)

    def _cleanup_old_exports(self, exports_dir: Path) -> None:
        cutoff = datetime.now(tz=timezone.utc) - timedelta(hours=24)
        for f in exports_dir.glob("*"):
            if f.is_file():
                mtime = datetime.fromtimestamp(f.stat().st_mtime, tz=timezone.utc)
                if mtime < cutoff:
                    try:
                        f.unlink()
                    except OSError:
                        pass
