import csv
import io
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from sqlalchemy import String, cast, func, or_, select
from sqlalchemy.orm import Session

from app.models.audit_log import AuditLog, derive_module


class AuditLogRepository:
    def __init__(self, db: Session) -> None:
        self.db = db

    # ------------------------------------------------------------------
    # Write
    # ------------------------------------------------------------------

    def log(
        self,
        action: str,
        *,
        user_id: uuid.UUID | None = None,
        ip_address: str | None = None,
        user_agent: str | None = None,
        module: str | None = None,
        entity_type: str | None = None,
        entity_id: str | None = None,
        entity_name: str | None = None,
        performed_by_name: str | None = None,
        old_values: dict | None = None,
        new_values: dict | None = None,
        remarks: str | None = None,
    ) -> AuditLog:
        record = AuditLog(
            action=action,
            module=module or derive_module(action),
            user_id=user_id,
            ip_address=ip_address,
            user_agent=user_agent,
            entity_type=entity_type,
            entity_id=str(entity_id) if entity_id is not None else None,
            entity_name=entity_name,
            performed_by_name=performed_by_name,
            old_values=old_values,
            new_values=new_values,
            remarks=remarks,
        )
        self.db.add(record)
        self.db.commit()
        return record

    # ------------------------------------------------------------------
    # Read
    # ------------------------------------------------------------------

    def get(self, log_id: uuid.UUID) -> AuditLog | None:
        return self.db.get(AuditLog, log_id)

    def list_paginated(
        self,
        *,
        page: int = 1,
        page_size: int = 25,
        search: str | None = None,
        module: str | None = None,
        action: str | None = None,
        entity_type: str | None = None,
        user_id: uuid.UUID | None = None,
        date_from: datetime | None = None,
        date_to: datetime | None = None,
        sort_by: str = "created_at",
        sort_order: str = "desc",
    ) -> dict[str, Any]:
        stmt = select(AuditLog)

        # Search across entity_name, action, performed_by_name
        if search:
            term = f"%{search}%"
            stmt = stmt.where(
                or_(
                    AuditLog.entity_name.ilike(term),
                    AuditLog.action.ilike(term),
                    AuditLog.performed_by_name.ilike(term),
                )
            )

        # Filters
        if module:
            stmt = stmt.where(AuditLog.module == module)
        if action:
            stmt = stmt.where(AuditLog.action == action)
        if entity_type:
            stmt = stmt.where(AuditLog.entity_type == entity_type)
        if user_id:
            stmt = stmt.where(AuditLog.user_id == user_id)
        if date_from:
            stmt = stmt.where(AuditLog.created_at >= date_from)
        if date_to:
            stmt = stmt.where(AuditLog.created_at <= date_to)

        # Total count
        count_stmt = select(func.count()).select_from(stmt.subquery())
        total = self.db.scalar(count_stmt) or 0

        # Sort
        sort_col_map = {
            "created_at": AuditLog.created_at,
            "module": AuditLog.module,
            "action": AuditLog.action,
            "performed_by_name": AuditLog.performed_by_name,
        }
        col = sort_col_map.get(sort_by, AuditLog.created_at)
        stmt = stmt.order_by(col.desc() if sort_order == "desc" else col.asc())

        # Pagination
        offset = (page - 1) * page_size
        stmt = stmt.offset(offset).limit(page_size)

        items = list(self.db.scalars(stmt).all())
        total_pages = max(1, -(-total // page_size))  # ceiling division

        return {
            "items": items,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": total_pages,
        }

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------

    def generate_export(
        self,
        *,
        filters: dict,
        fmt: str,
        storage_root: str,
    ) -> tuple[str, datetime]:
        """Generate a CSV or XLSX export file, return (filename, expires_at)."""
        # Collect all matching rows (no pagination for export)
        stmt = select(AuditLog)

        search = filters.get("search")
        if search:
            term = f"%{search}%"
            stmt = stmt.where(
                or_(
                    AuditLog.entity_name.ilike(term),
                    AuditLog.action.ilike(term),
                    AuditLog.performed_by_name.ilike(term),
                )
            )
        if filters.get("module"):
            stmt = stmt.where(AuditLog.module == filters["module"])
        if filters.get("action"):
            stmt = stmt.where(AuditLog.action == filters["action"])
        if filters.get("entity_type"):
            stmt = stmt.where(AuditLog.entity_type == filters["entity_type"])
        if filters.get("date_from"):
            stmt = stmt.where(AuditLog.created_at >= filters["date_from"])
        if filters.get("date_to"):
            stmt = stmt.where(AuditLog.created_at <= filters["date_to"])

        stmt = stmt.order_by(AuditLog.created_at.desc())
        rows = list(self.db.scalars(stmt).all())

        # Ensure export directory exists
        export_dir = Path(storage_root) / "exports"
        export_dir.mkdir(parents=True, exist_ok=True)

        ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"activity_{ts}.{fmt}"
        file_path = export_dir / filename
        expires_at = datetime.now(timezone.utc) + timedelta(hours=24)

        headers = [
            "Timestamp", "Module", "Action", "Entity Type", "Entity Name",
            "Entity ID", "Performed By", "IP Address", "User Agent",
            "Old Values", "New Values", "Remarks",
        ]

        def _row(r: AuditLog) -> list:
            return [
                r.created_at.isoformat() if r.created_at else "",
                r.module or "",
                r.action,
                r.entity_type or "",
                r.entity_name or "",
                r.entity_id or "",
                r.performed_by_name or "",
                r.ip_address or "",
                r.user_agent or "",
                str(r.old_values) if r.old_values else "",
                str(r.new_values) if r.new_values else "",
                r.remarks or "",
            ]

        if fmt == "csv":
            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerow(headers)
            for r in rows:
                writer.writerow(_row(r))
            file_path.write_text(buf.getvalue(), encoding="utf-8")

        else:  # xlsx
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Activity Logs"

            header_fill = PatternFill("solid", fgColor="1F4959")
            header_font = Font(bold=True, color="FFFFFF")

            ws.append(headers)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

            for r in rows:
                ws.append(_row(r))

            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

            wb.save(str(file_path))

        return filename, expires_at
